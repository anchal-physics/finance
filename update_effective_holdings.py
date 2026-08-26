#!/usr/bin/env python3
"""
update_effective_holdings.py
------------------------------------------------------------------------------
Look-through portfolio exposure.

The portfolio summary is now MULTI-ACCOUNT: GET_ALL_STOCK_SUMMARIES emits one
row per (account, ticker). For each Portfolio_* sheet this reads N (account),
O (ticker) and W (current value), decomposes every ETF position into its
underlying holdings (via yahooquery), and writes TWO surgical blocks on the
SAME sheet:

  SANKEY flows  → AN (Source) | AO (Value $) | AP (Target)
      A THREE-LEVEL graph, one row per edge:
        account → ticker      (that account's position value)
        ticker  → effective holding (look-through company / residual)
      Capuchin reads this block to draw the effective-holdings Sankey panel.
      All but the top EFF_TOP_TARGETS companies fold into an "Other holdings"
      node; each ETF's un-decomposed remainder is a "<ETF> — other holdings"
      target so dollars reconcile. (Account→ticker edges are never folded.)
  SUMMARY table → AR (Holding) | AS (Value $) | AT (% of portfolio, a FRACTION)
      The flat effective-holdings table (kept for reference).

Both blocks have a header row. Only those columns are cleared + rewritten;
nothing else is touched. A guard skips any sheet whose N1 isn't "Ticker".

Modes:
    python3 update_effective_holdings.py --dry-run     # print, write nothing
    python3 update_effective_holdings.py               # write into local Finance.xlsx
    python3 update_effective_holdings.py --online       # write to the Google Sheet

BOTH modes are surgical (XML-level for local, cell-range for online): only the
AM:AO and AQ:AS cells change — cached formula values, images, and every other
sheet/cell are left exactly as-is. Local mode writes a Finance.xlsx.bak first.
See SERVICE_ACCOUNT_SETUP.md for the one-time --online credential setup.

Requirements:  pip install -r requirements.txt
Online setup:  Google Cloud project w/ Sheets+Drive API, a Service Account JSON
               key saved as service_account.json (gitignored), the sheet shared
               with that service account (Editor), and FINANCE_SHEET_ID set.
Look-through note: yahooquery returns only each ETF's ~top-10 holdings, so the
un-listed remainder is booked to a "<ETF> — other holdings" line to keep the
dollar totals reconciled.
------------------------------------------------------------------------------
"""
import os
import sys
import shutil
import argparse
from collections import OrderedDict


def _load_local_env(path=None):
    """Load KEY=VALUE lines from a gitignored `.env` next to this script into
    os.environ WITHOUT overriding values already set (so real env / CI wins)."""
    path = path or os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.exists(path):
        return
    with open(path) as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


_load_local_env()

# ---- config (environment variables override these; see .env / .env.example) ----
LOCAL_FILE = os.environ.get("FINANCE_XLSX", os.path.join(os.path.dirname(__file__), "Finance.xlsx"))
SPREADSHEET_ID = os.environ.get("FINANCE_SHEET_ID", "PUT-SPREADSHEET-ID-HERE")
SERVICE_ACCOUNT_FILE = os.environ.get("GSPREAD_SA_FILE", "service_account.json")
SHEETS = os.environ.get("EFF_SHEETS", "Portfolio_AG,Portfolio_AA").split(",")

# Two output blocks (both 3 columns wide, with a header row):
#   SANKEY  : Source | Value | Target  — one row per flow edge (Capuchin reads
#             this to draw the effective-holdings Sankey panel).
#   SUMMARY : Holding | Value ($) | %   — the flat effective-holdings table.
SANKEY_START_COL = os.environ.get("EFF_SANKEY_COL", "AN")    # → AN, AO, AP
SUMMARY_START_COL = os.environ.get("EFF_SUMMARY_COL", "AR")  # → AR, AS, AT
try:
    TOP_TARGETS = int(os.environ.get("EFF_TOP_TARGETS") or 24)   # distinct right nodes; rest → "Other holdings"
except ValueError:
    TOP_TARGETS = 24

MAX_SCAN_ROW = 300          # portfolio positions never exceed this row
RESIDUAL_MIN_FRAC = 0.002   # ignore an ETF's residual if < 0.2% of the position
CASH_TICKERS = {"SPAXX", "FDRXX", "FZFXX", "SPRXX", "VMFXX"}  # money-market sweeps → not holdings


# ============================ shared helpers ============================
def col_to_num(col):
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def num_to_col(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def to_float(x):
    try:
        return float(str(x).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None


def fetch_lookthrough(tickers):
    """(holdings_by_ticker, names_by_ticker).
    holdings_by_ticker[t] = [{symbol,name,pct}] for ETFs, or None for a stock."""
    from yahooquery import Ticker
    tk = Ticker(tickers, asynchronous=False)
    hi = tk.fund_holding_info or {}
    pr = tk.price or {}
    holdings, names = {}, {}
    for t in tickers:
        info = hi.get(t)
        p = pr.get(t) if isinstance(pr.get(t), dict) else {}
        names[t] = p.get("longName") or p.get("shortName") or t
        hs = None
        if isinstance(info, dict) and isinstance(info.get("holdings"), list) and info["holdings"]:
            raw = info["holdings"]
            tot = sum((h.get("holdingPercent") or 0) for h in raw)
            scale = 0.01 if tot > 1.5 else 1.0   # guard if percents come as 0-100
            hs = []
            for h in raw:
                sym = (h.get("symbol") or "").strip()
                nm = (h.get("holdingName") or sym or "").strip()
                hs.append({"symbol": sym, "name": nm, "pct": (h.get("holdingPercent") or 0) * scale})
        holdings[t] = hs
    return holdings, names


def build_effective(positions, holdings, names):
    eff = OrderedDict()   # key -> {name, value}
    total = sum((p["value"] or 0) for p in positions if p["ticker"].upper() not in CASH_TICKERS)

    def add(key, name, val):
        if key not in eff:
            eff[key] = {"name": name, "value": 0.0}
        eff[key]["value"] += val

    for p in positions:
        t, val = p["ticker"], (p["value"] or 0)
        if val <= 0 or t.upper() in CASH_TICKERS:
            continue
        hs = holdings.get(t)
        if hs:  # ETF → distribute across underlyings; remainder → residual line
            allocated = 0.0
            for h in hs:
                key = h["symbol"] or h["name"]
                if not key:
                    continue
                add(key, h["name"] or key, val * h["pct"])
                allocated += h["pct"]
            resid = val * max(0.0, 1.0 - allocated)
            if resid > RESIDUAL_MIN_FRAC * val:
                add("{}::other".format(t), "{} — other holdings".format(t), resid)
        else:   # direct stock (or holdings unavailable) → itself
            add(t, names.get(t, t), val)

    rows = [{"name": v["name"], "value": v["value"],
             "pct": (v["value"] / total if total else 0.0)} for v in eff.values()]
    rows.sort(key=lambda r: -r["value"])
    return rows, total


def summary_table(rows):
    """Effective-holdings summary as a 2D list (header + rows): name / $ / %."""
    data = [["Effective Holding", "Value ($)", "% of Portfolio"]]
    for r in rows:
        data.append([r["name"], round(r["value"], 2), round(r["pct"], 6)])
    return data


def build_flows(positions, holdings, names, top_k):
    """Three-level Sankey edges as (source, value, target) tuples:
        LEVEL 1  account → ticker            (that account's position value)
        LEVEL 2  ticker  → effective holding (look-through company / residual)
    Level-2 edges are aggregated across accounts by (ticker, target); all but
    the top-`top_k` target companies fold into one 'Other holdings' node so the
    diagram stays readable. Level-1 (account→ticker) edges are never folded."""
    acct_edges = OrderedDict()   # (account, ticker) -> value   (level 1)
    edges = OrderedDict()        # (ticker, tkey)    -> value   (level 2)
    tname = {}                   # tkey -> display name
    tvalue = {}                  # tkey -> total value (for top-k ranking)

    def add(source, tkey, tnm, val):
        edges[(source, tkey)] = edges.get((source, tkey), 0.0) + val
        tname[tkey] = tnm
        tvalue[tkey] = tvalue.get(tkey, 0.0) + val

    for p in positions:
        t, val = p["ticker"], (p["value"] or 0)
        if val <= 0 or t.upper() in CASH_TICKERS:
            continue
        acct = (p.get("account") or "Unknown").strip() or "Unknown"
        acct_edges[(acct, t)] = acct_edges.get((acct, t), 0.0) + val
        hs = holdings.get(t)
        if hs:  # ETF → one edge per underlying + a residual edge
            allocated = 0.0
            for h in hs:
                key = h["symbol"] or h["name"]
                if not key:
                    continue
                add(t, key, h["name"] or key, val * h["pct"])
                allocated += h["pct"]
            resid = val * max(0.0, 1.0 - allocated)
            if resid > RESIDUAL_MIN_FRAC * val:
                add(t, "{}::other".format(t), "{} — other holdings".format(t), resid)
        else:   # direct stock → itself (skip if the company name == ticker to
                # avoid a self-loop; then the ticker node is simply terminal)
            comp = names.get(t, t)
            if comp and comp != t:
                add(t, "direct::{}".format(t), comp, val)

    keep = set(sorted(tvalue, key=lambda k: -tvalue[k])[:top_k])
    agg = OrderedDict()     # (ticker, target_label) -> value
    for (source, tkey), val in edges.items():
        label = tname[tkey] if tkey in keep else "Other holdings"
        agg[(source, label)] = agg.get((source, label), 0.0) + val

    # Level 1 first (account→ticker), each level largest-edge-first.
    flows = [(a, round(v, 2), t) for (a, t), v in
             sorted(acct_edges.items(), key=lambda e: (-e[1], e[0][0])) if v > 0]
    flows += [(s, round(v, 2), tl) for (s, tl), v in
              sorted(agg.items(), key=lambda e: (-e[1], e[0][0])) if v > 0]
    return flows


def flow_table(flows):
    """Sankey block as a 2D list (header + rows): Source / Value / Target."""
    data = [["Source", "Value", "Target"]]
    for s, v, t in flows:
        data.append([s, v, t])
    return data


QUIET = False   # set by --quiet: suppress dollar figures (public CI logs)


def print_preview(name, rows, total, flows):
    targets = {t for _, _, t in flows}
    sources = {s for s, _, _ in flows}
    if QUIET:
        # No dollar figures — safe for public Action logs.
        print("    {} effective holdings; sankey: {} flows, {} sources → {} targets".format(
            len(rows), len(flows), len(sources), len(targets)))
        return
    print("    portfolio total ${:,.2f}; {} effective holdings; "
          "sankey: {} flows, {} sources → {} targets".format(
              total, len(rows), len(flows), len(sources), len(targets)))
    for r in rows[:12]:
        print("      {:40}  ${:>12,.2f}  {:6.2f}%".format(r["name"][:40], r["value"], r["pct"] * 100))
    if len(rows) > 12:
        print("      … +{} more".format(len(rows) - 12))


# ============================ local (.xlsx) backend ============================
def read_positions_xlsx(ws):
    # Summary spill: N Account | O Ticker | P Total Shares | ... | W Value.
    out = []
    for row in range(2, MAX_SCAN_ROW + 1):
        acct = ws["N{}".format(row)].value
        tk = ws["O{}".format(row)].value
        if tk is None or str(tk).strip() == "":
            break  # the spill is contiguous; stop at the first blank ticker
        tk = str(tk).strip()
        if "#REF" in tk:
            continue
        out.append({"account": str(acct or "").strip(),
                    "ticker": tk,
                    "shares": to_float(ws["P{}".format(row)].value),
                    "value": to_float(ws["W{}".format(row)].value)})
    return out


SS_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _sheet_xml_path(zf, sheet_name):
    import re
    wb = zf.read("xl/workbook.xml").decode("utf-8")
    m = re.search(r'<sheet[^>]*name="{}"[^>]*r:id="(rId\d+)"'.format(re.escape(sheet_name)), wb)
    if not m:
        return None
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    m2 = re.search(r'Id="{}"[^>]*Target="([^"]*)"'.format(m.group(1)), rels)
    tgt = m2.group(1)
    return tgt if tgt.startswith("xl/") else "xl/" + tgt.lstrip("/")


def _col_of(ref):
    import re
    return re.match(r"([A-Z]+)", ref).group(1)


def write_blocks_xlsx(path, sheet_name, blocks):
    """SURGICALLY write one or more column blocks onto ONE sheet in a single pass.
    `blocks` is a list of (start_col, data) where data is a 2D list (header +
    rows). Each block occupies consecutive columns; its columns are set for
    rows 1..len(data) and cleared below. Every other byte of the workbook —
    other sheets, cached formula values, images, styles, shared strings — is
    left exactly as-is (only this sheet's XML is re-emitted, only these cells
    change)."""
    import zipfile, xml.etree.ElementTree as ET
    specs = []   # (c0, cols_set, data)
    for start_col, data in blocks:
        c0 = col_to_num(start_col)
        width = max((len(line) for line in data), default=1)
        specs.append((c0, {num_to_col(c0 + i) for i in range(width)}, data))

    with zipfile.ZipFile(path, "r") as zf:
        sheet_path = _sheet_xml_path(zf, sheet_name)
        if not sheet_path:
            raise RuntimeError("sheet {} not found in {}".format(sheet_name, path))
        names = zf.namelist()
        contents = {n: zf.read(n) for n in names}

    for pfx, uri in [("", SS_NS),
                     ("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships"),
                     ("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006"),
                     ("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")]:
        ET.register_namespace(pfx, uri)
    root = ET.fromstring(contents[sheet_path])
    sd = root.find("{%s}sheetData" % SS_NS)
    rows_by_r = {int(r.get("r")): r for r in sd.findall("{%s}row" % SS_NS)}

    def cell(colnum, rownum, value):
        c = ET.Element("{%s}c" % SS_NS)
        c.set("r", num_to_col(colnum) + str(rownum))
        if isinstance(value, str):
            c.set("t", "inlineStr")
            ET.SubElement(ET.SubElement(c, "{%s}is" % SS_NS), "{%s}t" % SS_NS).text = value
        else:
            ET.SubElement(c, "{%s}v" % SS_NS).text = repr(value) if isinstance(value, float) else str(value)
        return c

    def ensure_row(rownum):
        row = rows_by_r.get(rownum)
        if row is None:
            row = ET.SubElement(sd, "{%s}row" % SS_NS)
            row.set("r", str(rownum))
            rows_by_r[rownum] = row
        return row

    def strip(row, cols):
        for c in list(row):
            if c.get("r") and _col_of(c.get("r")) in cols:
                row.remove(c)

    for c0, cols, data in specs:
        for i, line in enumerate(data):                 # write header + rows
            row = ensure_row(1 + i)
            strip(row, cols)
            for dc, val in enumerate(line):
                row.append(cell(c0 + dc, 1 + i, val))
        for rownum, row in rows_by_r.items():           # clear this block below its data
            if rownum > len(data):
                strip(row, cols)

    for row in rows_by_r.values():                      # keep each row's cells column-ordered
        cells = sorted(list(row), key=lambda c: col_to_num(_col_of(c.get("r"))))
        for c in list(row):
            row.remove(c)
        for c in cells:
            row.append(c)
    rows_sorted = sorted(sd.findall("{%s}row" % SS_NS), key=lambda r: int(r.get("r")))
    for r in list(sd):
        sd.remove(r)
    for r in rows_sorted:
        sd.append(r)

    contents[sheet_path] = ET.tostring(root, encoding="UTF-8", xml_declaration=True)
    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
        for n in names:
            zf.writestr(n, contents[n])
    shutil.move(tmp, path)


def run_local(path, sheets, dry_run):
    from openpyxl import load_workbook
    if not os.path.exists(path):
        sys.exit("Local workbook not found: {}".format(path))
    print("Reading {}".format(path))
    wb_r = load_workbook(path, data_only=True, read_only=True)   # cached values only; never saved
    results = {}
    for name in sheets:
        name = name.strip()
        print("▸ {}".format(name))
        if name not in wb_r.sheetnames:
            print("    sheet not found — skipping")
            continue
        ws = wb_r[name]
        if str(ws["N1"].value or "").strip() != "Account":
            print("    N1 is not 'Account' → not a portfolio-summary layout; skipping.")
            continue
        positions = read_positions_xlsx(ws)
        tickers = sorted({p["ticker"] for p in positions})
        if not tickers:
            print("    no positions found")
            continue
        print("    {} positions; fetching look-through for {} tickers…".format(len(positions), len(tickers)))
        holdings, names = fetch_lookthrough(tickers)
        rows, total = build_effective(positions, holdings, names)
        flows = build_flows(positions, holdings, names, TOP_TARGETS)
        print_preview(name, rows, total, flows)
        results[name] = (rows, flows)
    wb_r.close()

    if dry_run:
        print("[dry-run] nothing written.")
        return
    if not results:
        print("Nothing to write.")
        return

    bak = path + ".bak"
    shutil.copy2(path, bak)
    print("Backup → {}".format(bak))
    for name, (rows, flows) in results.items():
        write_blocks_xlsx(path, name, [
            (SANKEY_START_COL, flow_table(flows)),      # Source | Value | Target
            (SUMMARY_START_COL, summary_table(rows)),   # Holding | Value | %
        ])
    print("Wrote sankey {}:{} + summary {}:{} to [{}] in {}".format(
        SANKEY_START_COL, num_to_col(col_to_num(SANKEY_START_COL) + 2),
        SUMMARY_START_COL, num_to_col(col_to_num(SUMMARY_START_COL) + 2),
        ", ".join(results), path))


# ============================ online (Google Sheet) backend ============================
def run_online(sheets, dry_run):
    import gspread
    if "PUT-SPREADSHEET-ID" in SPREADSHEET_ID:
        sys.exit("Set FINANCE_SHEET_ID (or edit SPREADSHEET_ID) to the Google Sheet's id.")
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        sys.exit("Service-account key not found: {}. See setup notes at the top of this file.".format(SERVICE_ACCOUNT_FILE))
    gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)
    sh = gc.open_by_key(SPREADSHEET_ID)

    def write_block_online(ws, start_col, data):
        c0 = col_to_num(start_col)
        cN, cP = num_to_col(c0), num_to_col(c0 + max((len(l) for l in data), default=1) - 1)
        ws.batch_clear(["{}1:{}{}".format(cN, cP, MAX_SCAN_ROW * 4)])
        ws.batch_update([{"range": "{}1:{}{}".format(cN, cP, len(data)), "values": data}])

    for name in sheets:
        name = name.strip()
        print("▸ {}".format(name))
        try:
            ws = sh.worksheet(name)
        except gspread.WorksheetNotFound:
            print("    sheet not found — skipping")
            continue
        if (ws.acell("N1").value or "").strip() != "Account":
            print("    N1 is not 'Account' → skipping (won't clobber data).")
            continue
        # Spill: N Account | O Ticker | P Shares | Q | R | S | T | U | V | W Value
        positions = [{"account": (r[0] or "").strip(),
                      "ticker": r[1].strip(),
                      "shares": to_float(r[2]) if len(r) > 2 else None,
                      "value": to_float(r[9]) if len(r) > 9 else None}
                     for r in ws.get("N2:W{}".format(MAX_SCAN_ROW))
                     if r and len(r) > 1 and r[1] and "#REF" not in r[1]]
        tickers = sorted({p["ticker"] for p in positions})
        if not tickers:
            print("    no positions found")
            continue
        print("    {} positions; fetching look-through for {} tickers…".format(len(positions), len(tickers)))
        holdings, names = fetch_lookthrough(tickers)
        rows, total = build_effective(positions, holdings, names)
        flows = build_flows(positions, holdings, names, TOP_TARGETS)
        print_preview(name, rows, total, flows)
        if dry_run:
            continue
        write_block_online(ws, SANKEY_START_COL, flow_table(flows))
        write_block_online(ws, SUMMARY_START_COL, summary_table(rows))
        print("    wrote sankey {} flows + summary {} holdings".format(len(flows), len(rows)))


def main():
    ap = argparse.ArgumentParser(description="Write effective look-through holdings to Portfolio_* sheets.")
    ap.add_argument("--online", action="store_true", help="update the Google Sheet (default: local Finance.xlsx)")
    ap.add_argument("--dry-run", action="store_true", help="print, do not write")
    ap.add_argument("--file", default=LOCAL_FILE, help="local .xlsx path (local mode)")
    ap.add_argument("--sheets", nargs="*", default=SHEETS, help="sheet names to update")
    ap.add_argument("--quiet", action="store_true",
                    help="suppress dollar figures in output (for public CI logs)")
    args = ap.parse_args()
    global QUIET
    QUIET = args.quiet
    if args.online:
        run_online(args.sheets, args.dry_run)
    else:
        run_local(args.file, args.sheets, args.dry_run)


if __name__ == "__main__":
    main()
